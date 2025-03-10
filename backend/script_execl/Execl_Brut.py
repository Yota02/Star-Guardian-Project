import os
import pandas as pd
import re
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def extract_data_from_txt(file_path):
    data = {}
    with open(file_path, 'r') as file:
        lines = file.readlines()
        for line in lines:
            if line.strip():
                # Enlever les unités entre crochets
                line = re.sub(r'\[.*?\]', '', line).strip()
                
                parts = line.split('=')
                if len(parts) == 2:
                    key = parts[0].strip()
                    value = parts[1].strip()
                    data[key] = value
    return data

def generer_execl_avec_toute_les_donnees(directory_path, output_file):
    all_data = []

    for file_name in os.listdir(directory_path):
        file_path = os.path.join(directory_path, file_name)

        if file_name.endswith(".txt") and os.path.isfile(file_path):
            file_data = extract_data_from_txt(file_path)
            all_data.append(file_data)

    # Créer un DataFrame avec les données extraites
    df = pd.DataFrame(all_data)
    
    # Si TCA n'existe pas, trier par CREATION_DATE
    if 'CREATION_DATE' in df.columns:
        df['CREATION_DATE'] = pd.to_datetime(df['CREATION_DATE'])
        df = df.sort_values(by='CREATION_DATE')
        # Reconvertir en string si nécessaire
        df['CREATION_DATE'] = df['CREATION_DATE'].dt.strftime('%Y-%m-%dT%H:%M:%S.%f')
    
    # Créer un nouveau classeur ou charger l'existant
    try:
        wb = load_workbook(output_file)
    except FileNotFoundError:
        wb = Workbook()

    # Supprimer la feuille 'TOUS' si elle existe déjà
    if 'TOUS' in wb.sheetnames:
        del wb['TOUS']
    
    # Créer une nouvelle feuille 'TOUS'
    ws = wb.create_sheet('TOUS')

    # Ajouter l'en-tête
    for col_num, column_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_num, value=column_name)

    # Ajouter les données
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)
        
    # Sauvegarder le fichier
    wb.save(output_file)
    

