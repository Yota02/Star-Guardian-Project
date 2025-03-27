import os
import re
import shutil
from openpyxl import load_workbook


from backend.script_execl import Execl_Brut
from backend.script_extraction import AgeAnalyzer, Conjonction, Country, Dates_, Distance_Miss, Inclination, Maneuvrable, Object_type, Probabilite

import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

class SatelliteDataProcessor:
    """
    Classe pour traiter les données de satellites et générer des statistiques 
    dans un fichier Excel.
    """
    
    def __init__(self, dossier=None, chemin_sortie=None):
        """
        Initialise le processeur de données satellites.
        
        Args:
            dossier (str, optional): Chemin du dossier contenant les fichiers à analyser.
            chemin_sortie (str, optional): Chemin du fichier Excel de sortie.
        """
        self.dossier = dossier
        self.chemin_sortie = chemin_sortie
        
        self.wb = None
        self.ws = None
        
        self.chemin_modele = "sortie.xlsx"
        self.format_type = "excel"  # Format par défaut
            
        # Initialisation des analyseurs
        self.conjunction_analyzer = None
        self.country_analyzer = None
        self.date_analyzer = None
        self.inclination_analyzer = None
        self.maneuvrable_analyzer = None
        self.object_type_analyzer = None
        self.probability_analyzer = None
        self.miss_distance_analyzer = None
        self.satelliteAgeAnalyzer = None
        
        if dossier and chemin_sortie:
            self.initialize_analyzers()
    
    
    def setCheminModel(self, chemin_modele):
        self.chemin_modele = chemin_modele
    
    def getCheminModel(self):
        return self.chemin_modele
    
    def set_wb(self):
        """
        Définit le classeur (Workbook) à partir du chemin de sortie.
        """
        
        if not self.chemin_sortie:
            print("Chemin de sortie non spécifié.")
            return
        
        """ if not os.path.exists(self.chemin_sortie):
            print(f"Le fichier {self.chemin_sortie} n'existe pas.")
            return """
        
        try:
            self.wb = load_workbook(self.chemin_sortie)
            if 'STATISTIQUES' not in self.wb.sheetnames:
                self.ws = self.wb.create_sheet('STATISTIQUES')
            else:
                self.ws = self.wb['STATISTIQUES']
        except Exception as e:
            print(f"Erreur lors du chargement du fichier Excel : {e}")
            self.wb = None
            
    def set_dossier(self, dossier):
        """
        Définit le dossier à analyser.
        
        Args:
            dossier (str): Chemin du dossier contenant les fichiers à analyser.
        """
        self.dossier = dossier
    
    def set_chemin_sortie(self, chemin_sortie):
        """
        Définit le chemin du fichier Excel de sortie.
        
        Args:
            chemin_sortie (str): Chemin du fichier Excel de sortie.
        """
        self.chemin_sortie = chemin_sortie
        
        # Si le workbook est déjà chargé, le fermer avant de réinitialiser
        if self.wb:
            try:
                self.wb.close()
            except:
                pass  # Ignorer les erreurs à la fermeture
            self.wb = None
            self.ws = None
        
        # Réinitialiser tous les analyseurs à None
        self.conjunction_analyzer = None
        self.country_analyzer = None
        self.date_analyzer = None
        self.satelliteAgeAnalyzer = None
        self.inclination_analyzer = None
        self.maneuvrable_analyzer = None
        self.object_type_analyzer = None
        self.probability_analyzer = None
        self.miss_distance_analyzer = None
        
        # Initialiser à nouveau les analyseurs
        if self.dossier and self.chemin_sortie:
            try:
                self.initialize_analyzers()
                if not self.conjunction_analyzer:
                    print("Warning: L'analyzeur de conjonctions n'a pas été initialisé correctement.")
            except Exception as e:
                print(f"Erreur lors de l'initialisation des analyseurs: {e}")
        
    def initialize_analyzers(self):
        """
        Initialise tous les analyseurs avec le dossier spécifié.
        """        
        
        if not self.chemin_sortie:
            print("Chemin de sortie non spécifié")
            return

        self.set_wb()
        
        if not self.wb:
            print("Impossible de charger le fichier Excel.")
            return
             
        self.conjunction_analyzer = Conjonction.ConjunctionAnalyzer(self.dossier, self.chemin_sortie, self.ws, self.wb)
        self.country_analyzer = Country.CountryAnalyzer(self.dossier, self.chemin_sortie, self.ws, self.wb)
        self.date_analyzer = Dates_.DateAnalyzer(self.dossier, self.chemin_sortie, self.ws, self.wb)
        self.satelliteAgeAnalyzer = AgeAnalyzer.SatelliteAgeAnalyzer(self.dossier, self.chemin_sortie, self.ws, self.wb)
        self.inclination_analyzer = Inclination.InclinationAnalyzer(self.dossier, self.chemin_sortie, self.ws, self.wb)
        self.maneuvrable_analyzer = Maneuvrable.ManeuvrableAnalyzer(self.dossier, self.chemin_sortie, self.ws, self.wb)
        self.object_type_analyzer = Object_type.ObjectTypeAnalyzer(self.dossier, self.chemin_sortie, self.ws, self.wb)
        self.probability_analyzer = Probabilite.CollisionProbabilityAnalyzer(self.dossier, self.chemin_sortie, self.ws, self.wb)
        self.miss_distance_analyzer = Distance_Miss.MissDistanceAnalyzer(self.dossier, self.chemin_sortie, self.ws, self.wb)
       
    def nom_satellite(self):
        """
        Extrait le nom du premier satellite trouvé dans les fichiers du dossier.
        
        Returns:
            str or None: Nom du satellite ou None si non trouvé.
        """
        # Vérifie si le répertoire existe
        if not os.path.exists(self.dossier):
            print(f"Le répertoire {self.dossier} n'existe pas.")
            return None

        # Parcours des fichiers dans le répertoire
        for filename in os.listdir(self.dossier):
            filepath = os.path.join(self.dossier, filename)

            # Vérifie si c'est un fichier
            if os.path.isfile(filepath):
                try:
                    # Lecture du contenu du fichier
                    with open(filepath, 'r', encoding='utf-8') as file:
                        content = file.read()

                    # Recherche du premier OBJECT_NAME
                    match = re.search(r'OBJECT_NAME\s*=\s*(.+)', content)
                    if match:
                        # Nettoyage du nom pour retirer les caractères non souhaités
                        satellite_name = match.group(1).strip()
                        satellite_name = re.sub(r'[^\w\s-]', '', satellite_name)  # Retirer les caractères non désirés
                        return satellite_name
                    else:
                        print(f"OBJECT_NAME introuvable dans {filename}.")
                        continue  # Passer au fichier suivant

                except Exception as e:
                    print(f"Erreur lors du traitement du fichier {filename}: {e}")
                    continue  # Passer au fichier suivant en cas d'erreur

        return None
    
    def compter_fichiers(self):
        """
        Compte le nombre de fichiers dans le dossier spécifié.
        
        Returns:
            int: Nombre de fichiers dans le dossier.
        """
        if not self.dossier:
            raise ValueError("Dossier non spécifié.")
            
        return len([f for f in os.listdir(self.dossier) if os.path.isfile(os.path.join(self.dossier, f))])
    
    def getConjonctionAnalyzer(self):
        return self.conjunction_analyzer
        
    def copier_modele_excel(self, chemin_modele):
        """
        Copie un modèle Excel et le renomme avec le format de date et version.
        
        Args:
            chemin_modele (str): Chemin complet du fichier modèle Excel.
        
        Returns:
            str or None: Chemin du nouveau fichier Excel créé, ou None si échec.
        """
        # Vérifier que le modèle existe
        if not os.path.exists(chemin_modele):
            print(f"Le modèle {chemin_modele} n'existe pas.")
            return None
        
        # Obtenir le nom du satellite
        nom_satellite = self.nom_satellite()
        
        # Dossier de sortie = dossier du modèle
        dossier_sortie = os.path.dirname(chemin_modele)
        
        # Générer le nom de fichier avec version
        nom_fichier_base = self.generer_nom_fichier_versionne(dossier_sortie, nom_satellite)
        
        # Ajouter l'extension
        nouveau_nom = f"{nom_fichier_base}.xlsx"
        
        # Construire le chemin complet du nouveau fichier
        chemin_nouveau_fichier = os.path.join(dossier_sortie, nouveau_nom)
        
        try:
            # Copier le fichier
            shutil.copy2(chemin_modele, chemin_nouveau_fichier)
            
            # Mettre à jour le chemin de sortie de l'instance
            self.set_chemin_sortie(chemin_nouveau_fichier)
            
            return chemin_nouveau_fichier
        
        except Exception as e:
            print(f"Erreur lors de la copie du modèle : {e}")
            return None
    
    def getSortie(self):
        """
        Obtient le chemin de sortie, avec l'extension correcte selon le format choisi.
        
        Returns:
            str: Chemin du fichier de sortie
        """
        # Si le format n'est pas Excel, ajuster l'extension
        if self.format_type != "excel" and self.chemin_sortie:
            if self.format_type == "calc" and self.chemin_sortie.endswith('.xlsx'):
                return self.chemin_sortie.replace('.xlsx', '.ods')
                
        return self.chemin_sortie
    
    def set_format(self, format_type):
        """
        Définit le format à utiliser pour le traitement des fichiers.
        
        Parameters:
        format_type (str): Le format à utiliser ('excel', 'calc', 'csv')
        """
        self.format_type = format_type
        
        # Adaptation du chargement et de la sauvegarde selon le format
        if hasattr(self, 'set_wb'):
            original_set_wb = self.set_wb
            
            def set_wb_wrapper():
                if self.format_type == "excel":
                    # Utiliser le code existant pour Excel
                    original_set_wb()
                    """ elif self.format_type == "calc":
                    # Chargement spécifique pour LibreOffice Calc
                    try:
                        import pyoo
                        # Démarrer LibreOffice en arrière-plan
                        desktop = pyoo.Desktop('localhost', 2002)
                        # Ouvrir le modèle
                        self.wb = desktop.open_spreadsheet(self.getCheminModel())
                    except Exception as e:
                        print(f"Erreur lors du chargement avec LibreOffice Calc: {e}")
                        print("Retour au format Excel.")
                        self.format_type = "excel"
                        original_set_wb() """
                else:
                    # Format non reconnu, utiliser Excel par défaut
                    print(f"Format {self.format_type} non reconnu. Utilisation d'Excel.")
                    self.format_type = "excel"
                    original_set_wb()
            
            # Remplacer la méthode set_wb par notre wrapper
            self.set_wb = set_wb_wrapper
            
    def generer_nom_fichier_versionne(self, base_dir, satellite_name=None):
        """
        Génère un nom de fichier au format 'YYYY-MM-DD-SatName-v0.0.X' en fonction du satellite et incrémente la version.
        
        Args:
            base_dir (str): Répertoire où le fichier sera créé
            satellite_name (str, optional): Nom du satellite à inclure dans le nom du fichier
            
        Returns:
            str: Nom du fichier généré
        """
        from datetime import datetime
        import os
        import re
        
        # Obtenir la date actuelle au format YYYY-MM-DD
        date_str = datetime.now().strftime('%Y-%m-%d')
        
        if not satellite_name:
            raise ValueError("Le nom du satellite doit être spécifié pour générer un fichier versionné.")
        
        # Format du préfixe
        safe_name = re.sub(r'[^\w\s-]', '', satellite_name)  # Assurez-vous que le nom soit valide
        prefix = f"{date_str}-{safe_name}"
        
        # Pattern pour détecter les fichiers avec la version du satellite (SatName_vX.Y.Z)
        pattern = fr"{date_str}-{safe_name}-v(\d+)\.(\d+)\.(\d+)"
        
        # Variables pour la dernière version
        max_major, max_minor, max_patch = 0, 0, 0
        
        # Vérifier les fichiers existants pour déterminer la dernière version du satellite
        if os.path.exists(base_dir):
            for file in os.listdir(base_dir):
                # Utiliser search pour trouver le pattern n'importe où dans le nom du fichier
                match = re.search(pattern, file)
                if match:
                    major, minor, patch = map(int, match.groups())
                    # Comparaison lexicographique des versions
                    if (major > max_major) or \
                    (major == max_major and minor > max_minor) or \
                    (major == max_major and minor == max_minor and patch > max_patch):
                        max_major, max_minor, max_patch = major, minor, patch
        
        # Incrémenter la version mineure pour créer un nouveau fichier (s'il n'y a pas de fichier, la version commence à 0.0.1)
        max_patch += 1  # Toujours incrémenter même si c'est la première version pour ce satellite
        version = f"v{max_major}.{max_minor}.{max_patch}"
        
        # Retourner le nom complet du fichier
        return f"{prefix}-{version}"
    
    def extract_data_from_txt_all(self, file_path):
        data = {}
        with open(file_path, 'r') as file:
            lines = file.readlines()
            for line in lines:
                if line.strip():
                    # Enlever les unités entre crochets
                    line = re.sub(r'\[.*?\]', '', line).strip()
                    
                    parts = line.split('=')
                    if len(parts) == 2:
                        key = parts[0].strip()
                        value = parts[1].strip()
                        data[key] = value
        return data

    def generer_execl_avec_toute_les_donnees(self):
        all_data = []

        for file_name in os.listdir(self.dossier):
            file_path = os.path.join(self.dossier, file_name)

            if file_name.endswith(".txt") and os.path.isfile(file_path):
                file_data = self.extract_data_from_txt_all(file_path)
                all_data.append(file_data)

        # Créer un DataFrame avec les données extraites
        df = pd.DataFrame(all_data)
        
        # Si TCA n'existe pas, trier par CREATION_DATE
        if 'CREATION_DATE' in df.columns:
            df['CREATION_DATE'] = pd.to_datetime(df['CREATION_DATE'])
            df = df.sort_values(by='CREATION_DATE')
            # Reconvertir en string si nécessaire
            df['CREATION_DATE'] = df['CREATION_DATE'].dt.strftime('%Y-%m-%dT%H:%M:%S.%f')
        
        # Créer un nouveau classeur ou charger l'existant

        # Supprimer la feuille 'TOUS' si elle existe déjà
        if 'TOUS' in self.wb.sheetnames:
            del self.wb['TOUS']
        
        # Créer une nouvelle feuille 'TOUS'
        self.ws = self.wb.create_sheet('TOUS')

        # Ajouter l'en-tête
        for col_num, column_name in enumerate(df.columns, 1):
            self.ws.cell(row=1, column=col_num, value=column_name)

        # Ajouter les données
        for r in dataframe_to_rows(df, index=False, header=False):
            self.ws.append(r)
            
        # Sauvegarder le fichier
        self.wb.save(self.chemin_sortie)
        
    def convert_to_format(self, source_path, target_path, format_type):
        """
        Convertit un fichier Excel vers le format ODS pour LibreOffice Calc.
        
        Args:
            source_path (str): Chemin du fichier Excel source
            target_path (str): Chemin du fichier ODS de destination
            format_type (str): Format cible ('calc' pour ODS)
            
        Returns:
            bool: True si la conversion a réussi, False sinon
        """
        try:
            if format_type == "calc":
                # Vérifier que l'extension du fichier cible est .ods
                if not target_path.endswith('.ods'):
                    target_path = target_path.rsplit('.', 1)[0] + '.ods'
                
                import pandas as pd
                
                # Lire toutes les feuilles Excel
                excel_file = pd.ExcelFile(source_path)
                sheet_names = excel_file.sheet_names
                
                # Créer un writer pour le format ODS
                with pd.ExcelWriter(target_path, engine='odf') as writer:
                    # Copier chaque feuille
                    for sheet_name in sheet_names:
                        df = pd.read_excel(excel_file, sheet_name=sheet_name)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Vérifier que le fichier a bien été créé
                if os.path.exists(target_path):
                    print(f"Conversion réussie vers {target_path}")
                    return True
                else:
                    print(f"Échec: Fichier {target_path} non créé")
                    return False
                    
            else:
                print(f"Format de conversion '{format_type}' non supporté.")
                return False
                
        except ImportError:
            print("Erreur: Pandas avec support ODF non disponible.")
            print("Installez les bibliothèques nécessaires avec: pip install pandas odfpy")
            return False
        except Exception as e:
            print(f"Erreur lors de la conversion: {e}")
            import traceback
            traceback.print_exc()
            return False    
        
    def executer_analyse(self, racine_projet=None):
        """
        Exécute l'analyse complète en utilisant le chemin de sortie déjà configuré.
        
        Args:
            racine_projet (str, optional): Chemin de la racine du projet. 
                                        Si None, utilise le répertoire courant.
        
        Returns:
            bool: True si l'analyse s'est bien déroulée, False sinon.
        """
        try:
            # Vérifier que le chemin de sortie est configuré
            if not self.chemin_sortie:
                print("Erreur: Chemin de sortie non configuré.")
                return False
                
            # Obtenir le nom du satellite
            nom_satellite = self.nom_satellite()
            
            # Si le format est "calc", ajuster l'extension du chemin de sortie
            original_chemin_sortie = self.chemin_sortie
            if self.format_type == "calc" and not self.chemin_sortie.endswith('.ods'):
                self.chemin_sortie = self.chemin_sortie.rsplit('.', 1)[0] + '.ods'
                
            # On doit toujours travailler avec Excel temporairement pour la génération
            temp_excel_path = original_chemin_sortie
            if self.format_type == "calc":
                temp_excel_path = original_chemin_sortie.rsplit('.', 1)[0] + '.xlsx'
            
            # Copier le fichier modèle vers le chemin temporaire Excel
            try:
                shutil.copy2(self.chemin_modele, temp_excel_path)
            except Exception as e:
                print(f"Erreur lors de la copie du modèle: {e}")
                return False
            
            # Charger le nouveau fichier Excel pour le traitement
            temp_chemin_sortie = self.chemin_sortie
            self.chemin_sortie = temp_excel_path
            self.set_wb()
            
            # S'assurer que les analyseurs sont correctement initialisés
            self.initialize_analyzers()
            
            # Vérifier que les analyseurs sont bien initialisés
            if not self.conjunction_analyzer:
                print("L'analyseur de conjonction n'est pas initialisé après réinitialisation.")
                return False
            
            # Générer toutes les données dans le fichier Excel temporaire
            self.generer_execl_avec_toute_les_donnees()
            
            # Vérifier que le classeur est chargé
            if not self.wb:
                print("Classeur non chargé.")
                return False
            
            # Traitement des conjonctions
            self.conjunction_analyzer.process_data()
            self.conjunction_analyzer.generer_excel_avec_donnees()
            
            # Traitement des dates
            all_dates = self.date_analyzer._collect_dates()
            min_date = self.date_analyzer.find_min_date(all_dates)
            max_date = self.date_analyzer.find_max_date(all_dates)
            
            if 'STATISTIQUES' not in self.wb.sheetnames:
                self.ws = self.wb.create_sheet('STATISTIQUES')
            else:
                self.ws = self.wb['STATISTIQUES']

            # Ajouter les informations à la feuille de statistiques avec vérification
            try:
                # Écriture dans A1
                self.ws['A1'] = nom_satellite if nom_satellite else "Non trouvé"
                
                # Écriture dans D3
                fichiers_count = self.compter_fichiers()
                self.ws['D3'] = fichiers_count if fichiers_count is not None else 0
                
                # Écriture dans D4
                nb_conjunctions = self.conjunction_analyzer.get_conjunction_count()
                self.ws['D4'] = nb_conjunctions if nb_conjunctions is not None else 0
                
                
                
                # Écriture dans D6 et D7
                self.ws['D6'] = min_date.strftime('%Y-%m-%d') if min_date else 'Aucune date trouvée'
                self.ws['D7'] = max_date.strftime('%Y-%m-%d') if max_date else 'Aucune date trouvée'
                
                # Sauvegarder immédiatement les modifications
                self.wb.save(temp_excel_path)
                
            except Exception as e:
                print(f"Erreur lors de l'écriture dans les cellules : {e}")
                
            self.country_analyzer.process_data()
            self.inclination_analyzer.process_data()
            
            self.satelliteAgeAnalyzer.process_data()
            
            # Générer tous les données brutes
            Execl_Brut.generer_execl_avec_toute_les_donnees(self.dossier, temp_excel_path)
            
            # Sauvegarder le fichier Excel temporaire
            self.wb.save(temp_excel_path)
            
            # Si un format autre que Excel est demandé, convertir le fichier
            if self.format_type != "excel":
                # Restaurer le chemin de sortie original (potentiellement avec extension .ods)
                self.chemin_sortie = temp_chemin_sortie
                
                if self.convert_to_format(temp_excel_path, self.chemin_sortie, self.format_type):
                    print(f"Conversion réussie vers le format {self.format_type}: {self.chemin_sortie}")
                    # Si conversion réussie, supprimer le fichier Excel temporaire
                    if os.path.exists(self.chemin_sortie) and os.path.exists(temp_excel_path):
                        os.remove(temp_excel_path)
                else:
                    print(f"Échec de la conversion vers {self.format_type}. Le fichier Excel est conservé.")
                    # En cas d'échec, restaurer le chemin vers le fichier Excel
                    self.chemin_sortie = temp_excel_path
            
            print(f"Analyse terminée. Fichier sauvegardé : {self.chemin_sortie}")
            return True
        
        except Exception as e:
            import traceback
            print(f"Erreur lors de l'exécution de l'analyse: {e}")
            print(traceback.format_exc())  # Affiche la trace complète pour débogage
            return False