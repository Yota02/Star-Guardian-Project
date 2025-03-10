import os
import re
from collections import Counter
from openpyxl import load_workbook

from backend.script_extraction.ScriptAnalyzerABS import BaseAnalyzer

class CountryAnalyzer(BaseAnalyzer):
    """
    Classe pour analyser les données satellite, extraire les pays des opérateurs
    et exporter les résultats dans un fichier Excel.
    """
    
    def __init__(self, input=None, output=None, ws=None, wb=None):
        """
        Initialise l'analyseur de données satellite.
        
        Args:
            database_path (str, optional): Chemin vers le fichier Excel contenant le mapping opérateur-pays.
                                         Si None, utilise le chemin par défaut.
        """

        self.database_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../../config/Country_2025-01-24.xlsx")
        super().__init__(input, output, ws, wb)
        self.operator_mapping = self._get_operator_country_mapping()
    
    def _get_operator_country_mapping(self):
        """
        Récupère le mapping entre opérateurs et pays depuis le fichier Excel
        
        Returns:
            dict: Dictionnaire {opérateur: pays}
        """
        try:
            workbook = load_workbook(self.database_path)
            if "OPERATOR" not in workbook.sheetnames:
                print("La feuille 'OPERATOR' n'existe pas dans la base de données.")
                return {}
                
            sheet = workbook["OPERATOR"]
            
            # Trouver l'index de la colonne "COUNTRY"
            country_col = None
            for cell in sheet[1]:
                if cell.value == "COUNTRY":
                    country_col = cell.column_letter
                    break
            
            if not country_col:
                print("Colonne 'COUNTRY' non trouvée")
                return {}
                
            # Créer le mapping
            mapping = {}
            for row in sheet.iter_rows(min_row=2):
                operator = row[0].value
                if operator and sheet[f"{country_col}{row[0].row}"].value:
                    mapping[operator] = sheet[f"{country_col}{row[0].row}"].value
                    
            return mapping
            
        except Exception as e:
            print(f"Erreur lors de la lecture de la base de données: {str(e)}")
            return {}
    
    def analyze_folder(self):
        """
        Analyse le dossier pour extraire les pays des opérateurs en évitant 
        les doublons basés sur INTERNATIONAL_DESIGNATOR
        
        Returns:
            list: Liste des pays trouvés
        """
        countries = []
        seen_designators = set()
        
        # Parcourir tous les fichiers .txt du dossier
        for filename in os.listdir(self.input):
            if filename.endswith('.txt'):
                file_path = os.path.join(self.input, filename)
                
                try:
                    with open(file_path, 'r') as file:
                        contenu = file.read()
                        
                        # Chercher toutes les sections qui contiennent OBJECT2
                        sections = contenu.split('OBJECT                             =')
                        for section in sections:
                            if section.startswith('OBJECT2'):
                                
                                # Récupérer le INTERNATIONAL_DESIGNATOR
                                designator_match = re.search(r'INTERNATIONAL_DESIGNATOR\s*=\s*(\S+)', section)
                                if designator_match:
                                    designator = designator_match.group(1).strip()
                                    
                                    # Vérifier si déjà vu
                                    if designator in seen_designators:
                                        continue  # Ignorer si déjà traité
                                    seen_designators.add(designator)
                                
                                # Chercher la ligne OPERATOR_ORGANIZATION
                                operator_match = re.search(r'OPERATOR_ORGANIZATION\s*=\s*(.+)', section)
                                if operator_match:
                                    operator = operator_match.group(1).strip()
                                    if operator and operator != "NONE":
                                        country = self.operator_mapping.get(operator, None)
                                        if country is not None:
                                            countries.append(country)
                                        
                except Exception as e:
                    print(f"Erreur lors de la lecture du fichier {filename}: {str(e)}")
        
        return countries
    
    def get_unique_countries(self, countries):
        """
        Obtient la liste des pays uniques.
        
        Args:
            countries (list): Liste de pays (peut contenir des doublons)
            
        Returns:
            list: Liste des pays uniques
        """
        return list(set(countries))
    
    def export_to_excel(self, countries):
        """
        Exporte les données des pays vers la feuille de calcul spécifiée.
        
        Args:
            countries (list): Liste des pays à exporter
            worksheet: Feuille de calcul openpyxl où exporter les données
        """
        # Calculer la distribution des pays
        counter = Counter(countries)
        
        try:
            # Nettoyer les anciennes données
            for i in range(3, self.ws.max_row + 1):
                self.ws[f'AE{i}'] = None
                self.ws[f'AF{i}'] = None
            
            # Trier par nombre de fichiers (ordre décroissant)
            sorted_countries = sorted(counter.items(), key=lambda x: x[1], reverse=True)

            # Écrire les nouvelles données à partir de la ligne 3
            for i, (country, count) in enumerate(sorted_countries, start=3):
                self.ws[f'AE{i}'] = country
                self.ws[f'AF{i}'] = count
                
        except Exception as e:
            print(f"Erreur lors de l'export Excel: {str(e)}")
            
    def process_data(self):
        """
        Traite les données du dossier et exporte les résultats dans le fichier Excel.
        
        Args:
            input (str): Chemin du dossier à analyser
            output_workbook: Classeur Excel où exporter les résultats
            worksheet_name (str): Nom de la feuille où exporter les résultats
            
        Returns:
            list: Liste des pays trouvés
        """
        # Analyser le dossier
        countries = self.analyze_folder()
        
        self.export_to_excel(countries)
       
        return countries