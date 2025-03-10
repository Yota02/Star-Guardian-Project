import os
import openpyxl
import re
from datetime import datetime
from backend.script_extraction.ScriptAnalyzerABS import BaseAnalyzer

class DateAnalyzer(BaseAnalyzer):
    """
    Classe pour extraire et analyser les dates de création à partir de fichiers texte
    et exporter les résultats vers un fichier Excel.
    """
    
    def __init__(self, input, output, ws, wb):
        """Initialise l'extracteur de dates."""
        super().__init__(input, output, ws, wb)
        
    def extract_dates(self, file_path):
        """
        Extrait les dates de création à partir d'un fichier texte.
        
        Args:
            file_path (str): Chemin vers le fichier à analyser
            
        Returns:
            list: Liste des dates trouvées au format 'YYYY-MM-DD'
        """
        try:
            with open(file_path, 'r') as file:
                content = file.read()
                # Recherche de la date associée à CREATION_DATE
                match = re.search(r'CREATION_DATE\s*=\s*(\d{4}-\d{2}-\d{2})', content)
                if match:
                    return [match.group(1)]  # Retourne une liste contenant la date trouvée
                else:
                    return []  # Si CREATION_DATE n'est pas trouvé, retourne une liste vide
        except Exception as e:
            print(f"Erreur lors de la lecture du fichier {file_path}: {str(e)}")
            return []
        
    def find_min_date(self, dates):
        """
        Trouve la date minimale (la plus ancienne) parmi toutes les dates fournies.
        
        Args:
            dates (list): Liste de dates au format datetime.
            
        Returns:
            datetime: La date la plus ancienne, ou None si aucune date n'est trouvée
        """
        if dates:
            return min(dates)
        return None

    def find_max_date(self, dates):
        """
        Trouve la date maximale (la plus récente) parmi toutes les dates fournies.
        
        Args:
            dates (list): Liste de dates au format datetime.
            
        Returns:
            datetime: La date la plus récente, ou None si aucune date n'est trouvée
        """
        if dates:
            return max(dates)
        return None
    
    def process_data(self):
        """
        Traite tous les fichiers et extrait les dates, puis génère un résumé.
        """
        all_dates = self._collect_dates()
        
        # Calculer les dates minimales et maximales
        min_date = self.find_min_date(all_dates)
        max_date = self.find_max_date(all_dates)
        
        # Affichage des résultats
        print(f"Date la plus ancienne : {min_date.strftime('%Y-%m-%d') if min_date else 'Aucune date trouvée'}")
        print(f"Date la plus récente : {max_date.strftime('%Y-%m-%d') if max_date else 'Aucune date trouvée'}")
        
        # Exporter les dates vers un fichier Excel
        self.process_directory()

    def _collect_dates(self):
        """
        Collecte toutes les dates valides à partir des fichiers du répertoire.
        
        Args:
            input_directory (str): Chemin vers le répertoire contenant les fichiers à analyser
            
        Returns:
            list: Liste des objets datetime correspondant aux dates trouvées
        """
        all_dates = []
        
        for filename in os.listdir(self.input):
            if filename.endswith('.txt'):
                file_path = os.path.join(self.input, filename)
                file_dates = self.extract_dates(file_path)
                
                for date in file_dates:
                    try:
                        # Tente de convertir chaque date au format '%Y-%m-%d'
                        parsed_date = datetime.strptime(date, '%Y-%m-%d')
                        all_dates.append(parsed_date)
                    except ValueError:
                        # Si la date est invalide, on l'ignore et passe à la suivante
                        print(f"Ignorer date invalide: {date}")
                        continue
        
        return all_dates
    
    def process_directory(self):
        """
        Traite tous les fichiers du répertoire et exporte les dates vers un fichier Excel.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Filename")
        ws.cell(row=1, column=2, value="Date")
        
        row = 2
        try:
            for filename in os.listdir(self.input):
                if filename.endswith('.txt'):
                    file_path = os.path.join(self.input, filename)
                    file_dates = self.extract_dates(file_path)
                    
                    for date in file_dates:
                        ws.cell(row=row, column=1, value=filename)
                        ws.cell(row=row, column=2, value=date)
                        row += 1
            
            wb.save(self.output)
            print(f"Données exportées avec succès vers {self.output}")
            
        except Exception as e:
            print(f"Erreur lors du traitement ou de l'export: {str(e)}")
    
    def generate_date_summary(self):
        """
        Génère un résumé des dates trouvées dans le répertoire.
        
        Args:
            input_directory (str): Chemin vers le répertoire contenant les fichiers à analyser
            
        Returns:
            dict: Dictionnaire contenant le résumé des dates
        """
        all_dates = self._collect_dates()
        
        if not all_dates:
            return {
                "total_files": 0,
                "files_with_dates": 0,
                "earliest_date": None,
                "latest_date": None,
                "date_range_days": None
            }
        
        # Compter les fichiers
        total_files = sum(1 for f in os.listdir(self.input) if f.endswith('.txt'))
        files_with_dates = sum(1 for f in os.listdir(self.input) 
                              if f.endswith('.txt') and self.extract_dates(os.path.join(self.input, f)))
        
        earliest_date = min(all_dates)
        latest_date = max(all_dates)
        date_range = (latest_date - earliest_date).days
        
        return {
            "total_files": total_files,
            "files_with_dates": files_with_dates,
            "earliest_date": earliest_date.strftime('%Y-%m-%d'),
            "latest_date": latest_date.strftime('%Y-%m-%d'),
            "date_range_days": date_range
        }
