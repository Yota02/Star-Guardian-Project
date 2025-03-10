import os
import re
import pandas as pd
from datetime import datetime
from typing import Dict, List, Set, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from backend.script_extraction.ScriptAnalyzerABS import BaseAnalyzer

class ConjunctionAnalyzer(BaseAnalyzer):
    def __init__(self, input, output, ws, wb):
        super().__init__(input, output, ws, wb)
        self.object_designator_files_map: Dict[str, list] = {}
        self.conjunctions: Dict[int, Set[str]] = {}
        self.all_data: List[Dict] = []

    def extract_data_from_txt(self, file_path: str) -> Dict:
        data = {}
        with open(file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()
            for line in lines:
                if line.strip():
                    line = re.sub(r'\[.*?\]', '', line).strip()
                    parts = line.split('=')
                    if len(parts) == 2:
                        key = parts[0].strip()
                        value = parts[1].strip()
                        data[key] = value
        return data

    def extract_object_designators(self) -> Dict[str, list]:
        self.object_designator_files_map.clear()
        self.all_data.clear()

        for filename in os.listdir(self.input):
            if filename.endswith('.txt'):
                file_path = os.path.join(self.input, filename)
                
                try:
                    file_data = self.extract_data_from_txt(file_path)
                    file_data['FILENAME'] = filename
                    self.all_data.append(file_data)

                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                    
                    object_designator_matches = re.finditer(r'OBJECT\s*=\s*OBJECT2.*?OBJECT_DESIGNATOR\s*=\s*(\d+)', content, re.DOTALL)
                    
                    for match in object_designator_matches:
                        object_designator = match.group(1).strip()
                        if not object_designator:
                            continue
                            
                        if object_designator in self.object_designator_files_map:
                            if filename not in self.object_designator_files_map[object_designator]:
                                self.object_designator_files_map[object_designator].append(filename)
                        else:
                            self.object_designator_files_map[object_designator] = [filename]
                        
                except Exception as e:
                    print(f"Erreur de lecture du fichier {filename}: {str(e)}")
        
        return self.object_designator_files_map or {}

    @staticmethod
    def extract_tca(file_content: str) -> Optional[str]:
        tca_match = re.search(r'CREATION_DATE\s*=\s*([0-9\-T:.]+)', file_content)
        return tca_match.group(1).strip() if tca_match else None

    @staticmethod
    def is_conjunction(date1: str, date2: str) -> bool:
        try:
            date_format = "%Y-%m-%dT%H:%M:%S.%f"
            dt1 = datetime.strptime(date1, date_format)
            dt2 = datetime.strptime(date2, date_format)
            
            deltaT = abs((dt1 - dt2).total_seconds())
            return deltaT <= 86400  # 24 hours in seconds
        except ValueError as e:
            print(f"Error parsing dates: {date1} or {date2}. Error: {str(e)}")
            return False

    def analyze_conjunctions(self) -> Dict[int, Set[str]]:
        self.conjunctions.clear()
        conjunction_id = 1

        for object_designator, files in self.object_designator_files_map.items():
            tca_list = []
            
            # Extract TCAs for files
            for file in files:
                file_path = os.path.join(self.input, file)
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                        tca = self.extract_tca(content)
                        if tca:
                            tca_list.append((file, tca))
                except Exception as e:
                    print(f"Error reading file {file}: {str(e)}")
            
            # If only one file, create a group for this file
            if len(tca_list) == 1:
                self.conjunctions[conjunction_id] = {tca_list[0][0]}
                conjunction_id += 1
                continue
            
            # Compare files
            for i in range(len(tca_list)):
                for j in range(i + 1, len(tca_list)):
                    file1, tca1 = tca_list[i]
                    file2, tca2 = tca_list[j]
                    
                    if self.is_conjunction(tca1, tca2):
                        # Look for existing groups containing either file
                        existing_group = None
                        for group_id, group in self.conjunctions.items():
                            if file1 in group or file2 in group:
                                existing_group = group_id
                                break
                        
                        if existing_group is not None:
                            self.conjunctions[existing_group].add(file1)
                            self.conjunctions[existing_group].add(file2)
                        else:
                            self.conjunctions[conjunction_id] = {file1, file2}
                            conjunction_id += 1
            
        return self.conjunctions or {}        
    
    def get_conjunction_count(self) -> int:
        """
        Obtient le nombre de groupes de conjonction.
        
        Returns:
            int: Nombre de groupes de conjonction
        """
        return len(self.conjunctions)

    def process_data(self):
        self.extract_object_designators()
        self.analyze_conjunctions()
    
    def generer_excel_avec_donnees(self):
        """
        Génère un fichier Excel avec seulement le premier fichier de chaque groupe de conjonction.
        Remplace les points par des virgules dans toutes les valeurs.
        """
        # Collecter les premiers fichiers de chaque groupe de conjonction
        first_files_in_conjunctions = set()
        for group in self.conjunctions.values():
            first_files_in_conjunctions.add(list(group)[0])

        # Filtrer les données pour n'inclure que les fichiers des groupes de conjonction
        filtered_data = [
            data for data in self.all_data 
            if data['FILENAME'] in first_files_in_conjunctions
        ]
        
        # Créer un DataFrame avec les données filtrées
        df = pd.DataFrame(filtered_data)
        
        # Fonction pour formater les nombres et remplacer les points par des virgules
        def formater_valeur(valeur):
            # Convertir en chaîne pour pouvoir remplacer les points
            valeur_str = str(valeur)
            
            # Remplacer les points par des virgules
            valeur_str = valeur_str.replace('.', ',')
            
            return valeur_str

        # Appliquer le formatage à toutes les colonnes
        for col in df.columns:
            df[col] = df[col].apply(formater_valeur)

        # Gestion spécifique de la colonne MISS_DISTANCE
        if 'MISS_DISTANCE' in df.columns:
            # Convertir la colonne MISS_DISTANCE en nombres
            df['MISS_DISTANCE'] = df['MISS_DISTANCE'].str.replace(',', '.').astype(float)
            
            
        if 'COLLISION_PROBABILITY' in df.columns:
            # Formatter la colonne COLLISION_PROBABILITY spécifiquement
            df['COLLISION_PROBABILITY'] = df['COLLISION_PROBABILITY'].str.replace(',', '.').astype(float)
            
        # Supprimer la feuille 'SHORTLIST' si elle existe déjà
        if 'SHORTLIST' in self.wb.sheetnames:
            del self.wb['SHORTLIST']
        
        # Créer une nouvelle feuille 'SHORTLIST'
        ws = self.wb.create_sheet('SHORTLIST')

        # Ajouter l'en-tête
        for col_num, column_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_num, value=column_name)

        # Ajouter les données
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)
        
        # Sauvegarder le fichier
        self.wb.save(self.output)